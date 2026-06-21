"""
05_build_excel.py
Generates a professional Excel workbook for the Lamination Store project.
Sheets:
  1. Dashboard        - KPI summary with Excel formulas
  2. Raw Data         - Full fact_sales table
  3. Monthly Revenue  - Pivot-style summary with chart
  4. Product Analysis - Revenue/Profit/Margin per product
  5. Customer Analysis- Top customers + segment split
  6. Region & Employee- Sales-rep and region performance
"""

import sqlite3
import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              numbers)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import DataPoint

# ── Colors ────────────────────────────────────────────────────────────
DARK_BLUE  = "1F3864"
MID_BLUE   = "2E5984"
LIGHT_BLUE = "BDD7EE"
ACCENT     = "E8964D"
GREEN      = "375623"
WHITE      = "FFFFFF"
LIGHT_GREY = "F2F2F2"
DARK_GREY  = "595959"
BORDER_CLR = "BFBFBF"

# ── Helpers ────────────────────────────────────────────────────────────
def hfill(color): return PatternFill("solid", start_color=color, fgColor=color)
def border(style="thin"):
    s = Side(style=style, color=BORDER_CLR)
    return Border(left=s, right=s, top=s, bottom=s)
def hdr_font(white=True, sz=11, bold=True):
    return Font(name="Arial", bold=bold, size=sz, color=WHITE if white else "000000")
def body_font(bold=False, sz=10, color="000000"):
    return Font(name="Arial", bold=bold, size=sz, color=color)
def money_fmt(): return '#,##0.00'
def pct_fmt():   return '0.0%'

def write_header_row(ws, row, cols, bg=MID_BLUE):
    for c, label in enumerate(cols, 1):
        cell = ws.cell(row=row, column=c, value=label)
        cell.font  = hdr_font()
        cell.fill  = hfill(bg)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border()

def auto_width(ws, min_w=10, max_w=35):
    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0 for c in col), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max(min_w, min(max_len + 2, max_w))

# ── Load data ──────────────────────────────────────────────────────────
db = "/home/claude/lamination_project/data/lamination_store.db"
conn = sqlite3.connect(db)

fact = pd.read_sql("""
    SELECT o.order_id, o.order_date, o.order_status, o.payment_method, o.processing_days,
           c.customer_name, c.customer_type, c.city, c.state,
           e.employee_name, e.region,
           p.product_name, p.category, p.size, p.thickness,
           oi.quantity, oi.unit_price, p.unit_cost,
           oi.discount_pct, oi.line_total,
           ROUND((oi.unit_price - p.unit_cost)*oi.quantity*(1-oi.discount_pct),2) AS profit
    FROM order_items oi
    JOIN orders o ON o.order_id = oi.order_id
    JOIN customers c ON c.customer_id = o.customer_id
    JOIN products p ON p.product_id = oi.product_id
    JOIN employees e ON e.employee_id = o.employee_id
    WHERE o.order_status <> 'Cancelled'
""", conn)

monthly = pd.read_sql("""
    SELECT strftime('%Y-%m', o.order_date) AS Month,
           COUNT(DISTINCT o.order_id) AS Orders,
           ROUND(SUM(oi.line_total),2) AS Revenue,
           ROUND(SUM((oi.unit_price-p.unit_cost)*oi.quantity*(1-oi.discount_pct)),2) AS Profit
    FROM orders o JOIN order_items oi ON oi.order_id=o.order_id
    JOIN products p ON p.product_id=oi.product_id
    WHERE o.order_status<>'Cancelled'
    GROUP BY Month ORDER BY Month
""", conn)

products_sum = pd.read_sql("""
    SELECT p.product_name AS Product, p.category AS Category,
           SUM(oi.quantity) AS Units_Sold,
           ROUND(SUM(oi.line_total),2) AS Revenue,
           ROUND(SUM((oi.unit_price-p.unit_cost)*oi.quantity*(1-oi.discount_pct)),2) AS Profit,
           ROUND(100.0*SUM((oi.unit_price-p.unit_cost)*oi.quantity*(1-oi.discount_pct))/SUM(oi.line_total),1) AS Margin_Pct
    FROM order_items oi JOIN orders o ON o.order_id=oi.order_id
    JOIN products p ON p.product_id=oi.product_id
    WHERE o.order_status<>'Cancelled'
    GROUP BY p.product_name, p.category ORDER BY Revenue DESC
""", conn)

customers_sum = pd.read_sql("""
    SELECT c.customer_name AS Customer, c.customer_type AS Type, c.city AS City,
           COUNT(DISTINCT o.order_id) AS Orders,
           ROUND(SUM(oi.line_total),2) AS Revenue
    FROM orders o JOIN order_items oi ON oi.order_id=o.order_id
    JOIN customers c ON c.customer_id=o.customer_id
    WHERE o.order_status<>'Cancelled'
    GROUP BY c.customer_name, c.customer_type, c.city ORDER BY Revenue DESC LIMIT 20
""", conn)

emp_sum = pd.read_sql("""
    SELECT e.employee_name AS Employee, e.region AS Region,
           COUNT(DISTINCT o.order_id) AS Orders,
           ROUND(SUM(oi.line_total),2) AS Revenue,
           ROUND(SUM((oi.unit_price-p.unit_cost)*oi.quantity*(1-oi.discount_pct)),2) AS Profit
    FROM orders o JOIN order_items oi ON oi.order_id=o.order_id
    JOIN employees e ON e.employee_id=o.employee_id
    JOIN products p ON p.product_id=oi.product_id
    WHERE o.order_status<>'Cancelled'
    GROUP BY e.employee_name, e.region ORDER BY Revenue DESC
""", conn)

conn.close()

# ── Workbook ──────────────────────────────────────────────────────────
wb = Workbook()
wb.remove(wb.active)   # remove default sheet

# ═══════════════════════════════════════════════════════════════════════
# SHEET 1 – DASHBOARD
# ═══════════════════════════════════════════════════════════════════════
dash = wb.create_sheet("Dashboard")
dash.sheet_view.showGridLines = False

# Title banner
dash.merge_cells("A1:H2")
dash["A1"] = "LAMINATION STORE — WHOLESALE & RETAIL SUPPLIER"
dash["A1"].font = Font(name="Arial", bold=True, size=18, color=WHITE)
dash["A1"].fill = hfill(DARK_BLUE)
dash["A1"].alignment = Alignment(horizontal="center", vertical="center")

dash.merge_cells("A3:H3")
dash["A3"] = "Sales Performance Dashboard  |  2024 – 2025"
dash["A3"].font = Font(name="Arial", size=11, italic=True, color=DARK_GREY)
dash["A3"].alignment = Alignment(horizontal="center")

dash.row_dimensions[1].height = 35
dash.row_dimensions[2].height = 5

# KPI cards (row 5-9)
kpi_labels = ["Total Revenue", "Total Profit", "Profit Margin %",
               "Total Orders", "Avg Order Value", "Return Rate %"]
kpi_values = [
    "=RawData!S2",   # formula refs to helper column we'll add
    "=RawData!T2",
    "=RawData!U2",
    "=RawData!V2",
    "=RawData!W2",
    "=RawData!X2",
]

# Place KPI boxes at columns A,B,C,D,E,F (two rows of 3)
positions = [(5,"A"),(5,"C"),(5,"E"),(8,"A"),(8,"C"),(8,"E")]
for (row, col), label, val in zip(positions, kpi_labels, kpi_values):
    end_col = chr(ord(col)+1)
    dash.merge_cells(f"{col}{row}:{end_col}{row}")
    lbl_cell = dash[f"{col}{row}"]
    lbl_cell.value = label
    lbl_cell.font = Font(name="Arial", bold=True, size=10, color=WHITE)
    lbl_cell.fill = hfill(MID_BLUE)
    lbl_cell.alignment = Alignment(horizontal="center")

    dash.merge_cells(f"{col}{row+1}:{end_col}{row+1}")
    val_cell = dash[f"{col}{row+1}"]
    val_cell.value = val
    val_cell.font = Font(name="Arial", bold=True, size=14, color=DARK_BLUE)
    val_cell.fill = hfill(LIGHT_BLUE)
    val_cell.alignment = Alignment(horizontal="center", vertical="center")
    dash.row_dimensions[row+1].height = 28

# Section header
dash.merge_cells("A11:H11")
dash["A11"] = "Key Insights"
dash["A11"].font = Font(name="Arial", bold=True, size=12, color=WHITE)
dash["A11"].fill = hfill(ACCENT)
dash["A11"].alignment = Alignment(horizontal="left", indent=1)

insights = [
    "• Wholesale customers drive ~86% of total revenue despite being only 30% of customer base.",
    "• Cold Lamination Rolls and Thermal Lamination Rolls are the top revenue-generating product lines.",
    "• Film products have the highest profit margin (31.3%) among all categories.",
    "• Seasonal peaks observed in April–May and October–November (back-to-school & festive season).",
    "• 84.3% order delivery success rate; 9.1% returns require investigation.",
    "• Top 5 cities (Chandigarh, Bengaluru, Ahmedabad, Chennai, Nagpur) account for majority of revenue.",
]
for i, text in enumerate(insights, 12):
    dash[f"A{i}"] = text
    dash[f"A{i}"].font = body_font(sz=10)
    dash.merge_cells(f"A{i}:H{i}")

for col in ["A","B","C","D","E","F","G","H"]:
    dash.column_dimensions[col].width = 18

# ═══════════════════════════════════════════════════════════════════════
# SHEET 2 – RAW DATA
# ═══════════════════════════════════════════════════════════════════════
raw = wb.create_sheet("RawData")
cols = list(fact.columns)
write_header_row(raw, 1, cols)

for r_idx, row in fact.iterrows():
    for c_idx, val in enumerate(row, 1):
        cell = raw.cell(row=r_idx+2, column=c_idx, value=val)
        cell.font = body_font(sz=9)
        cell.border = border()
        if cols[c_idx-1] in ("line_total","profit","unit_price","unit_cost"):
            cell.number_format = money_fmt()
        elif cols[c_idx-1] == "discount_pct":
            cell.number_format = pct_fmt()

# Helper KPI cells in S2:X2 referenced by Dashboard
n = len(fact) + 1  # last data row in RawData (1-indexed, +1 for header)
raw["S2"] = f"=ROUND(SUM(T3:T{n+1}),2)"   # placeholder; we use named columns below
# Actually wire to line_total (col index 20 = T) and profit (col 21 = U)
raw["S2"] = f"=ROUND(SUM(T3:T{len(fact)+2}),2)"
raw["T2"] = f"=ROUND(SUM(U3:U{len(fact)+2}),2)"
raw["U2"] = f"=ROUND(T2/S2,3)"
raw["V2"] = f"=COUNTA(A3:A{len(fact)+2})-SUMPRODUCT((A3:A{len(fact)+2}=A2:A{len(fact)+1})*1)"
raw["W2"] = f"=ROUND(S2/V2,2)"
raw["X2"] = "0.091"  # static from SQL; could be dynamic with COUNTIF

# Freeze header row
raw.freeze_panes = "A2"
auto_width(raw, min_w=12, max_w=28)

# ═══════════════════════════════════════════════════════════════════════
# SHEET 3 – MONTHLY REVENUE
# ═══════════════════════════════════════════════════════════════════════
mws = wb.create_sheet("Monthly Revenue")
mws.sheet_view.showGridLines = False

mws.merge_cells("A1:E1")
mws["A1"] = "Monthly Revenue & Profit Summary"
mws["A1"].font = hdr_font(sz=13)
mws["A1"].fill = hfill(DARK_BLUE)
mws["A1"].alignment = Alignment(horizontal="center", vertical="center")
mws.row_dimensions[1].height = 28

write_header_row(mws, 2, ["Month","Orders","Revenue (Rs.)","Profit (Rs.)","Margin %"], bg=MID_BLUE)

for r, row in monthly.iterrows():
    dr = r + 3
    mws.cell(dr, 1, row["Month"]).font = body_font()
    mws.cell(dr, 2, row["Orders"]).font = body_font()

    rev_cell  = mws.cell(dr, 3, row["Revenue"])
    pft_cell  = mws.cell(dr, 4, row["Profit"])
    mrg_cell  = mws.cell(dr, 5)
    mrg_cell.value = f"=ROUND(D{dr}/C{dr},3)"

    for cell in [rev_cell, pft_cell]:
        cell.number_format = '#,##0'
        cell.font = body_font()
    mrg_cell.number_format = pct_fmt()
    mrg_cell.font = body_font()

    fill = hfill(LIGHT_GREY) if r % 2 == 0 else hfill(WHITE)
    for c in range(1, 6):
        mws.cell(dr, c).fill = fill
        mws.cell(dr, c).border = border()

# Totals row
last = len(monthly) + 2
tr = last + 1
mws.cell(tr, 1, "TOTAL").font = hdr_font(white=False, sz=10)
mws.cell(tr, 1).fill = hfill(LIGHT_BLUE)
for c_offset, col in [(1,"B"),(2,"C"),(3,"D")]:
    cell = mws.cell(tr, c_offset+1, f"=SUM({col}3:{col}{last})")
    cell.font = hdr_font(white=False, sz=10)
    cell.fill = hfill(LIGHT_BLUE)
    if c_offset >= 2:
        cell.number_format = '#,##0'
    cell.border = border()
mws.cell(tr, 5, f"=ROUND(D{tr}/C{tr},3)").number_format = pct_fmt()
mws.cell(tr, 5).fill = hfill(LIGHT_BLUE)

# Line Chart – Revenue
chart = LineChart()
chart.title = "Monthly Revenue Trend"
chart.style = 10
chart.y_axis.title = "Revenue (Rs.)"
chart.x_axis.title = "Month"
chart.height = 12
chart.width  = 20

data_ref = Reference(mws, min_col=3, min_row=2, max_row=len(monthly)+2)
cats_ref = Reference(mws, min_col=1, min_row=3, max_row=len(monthly)+2)
chart.add_data(data_ref, titles_from_data=True)
chart.set_categories(cats_ref)
chart.series[0].graphicalProperties.line.solidFill = MID_BLUE
chart.series[0].graphicalProperties.line.width     = 20000

mws.add_chart(chart, "G3")

for col, w in [("A",12),("B",9),("C",16),("D",16),("E",12)]:
    mws.column_dimensions[col].width = w
mws.freeze_panes = "A3"

# ═══════════════════════════════════════════════════════════════════════
# SHEET 4 – PRODUCT ANALYSIS
# ═══════════════════════════════════════════════════════════════════════
pws = wb.create_sheet("Product Analysis")
pws.sheet_view.showGridLines = False

pws.merge_cells("A1:F1")
pws["A1"] = "Product Revenue, Profit & Margin Analysis"
pws["A1"].font = hdr_font(sz=13)
pws["A1"].fill = hfill(DARK_BLUE)
pws["A1"].alignment = Alignment(horizontal="center", vertical="center")
pws.row_dimensions[1].height = 28

hdr_cols = ["Product","Category","Units Sold","Revenue (Rs.)","Profit (Rs.)","Margin %"]
write_header_row(pws, 2, hdr_cols)

for r, row in products_sum.iterrows():
    dr = r + 3
    pws.cell(dr, 1, row["Product"]).font = body_font(sz=9)
    pws.cell(dr, 2, row["Category"]).font = body_font(sz=9)
    pws.cell(dr, 3, row["Units_Sold"]).font = body_font(sz=9)
    rv = pws.cell(dr, 4, row["Revenue"])
    pf = pws.cell(dr, 5, row["Profit"])
    mg = pws.cell(dr, 6)
    mg.value = f"=ROUND(E{dr}/D{dr},3)"
    rv.number_format = '#,##0'
    pf.number_format = '#,##0'
    mg.number_format = pct_fmt()
    fill = hfill(LIGHT_GREY) if r % 2 == 0 else hfill(WHITE)
    for c in range(1,7):
        pws.cell(dr,c).fill = fill
        pws.cell(dr,c).border = border()
        pws.cell(dr,c).font = body_font(sz=9)

# Bar chart – Top 10 product revenue
chart2 = BarChart()
chart2.type  = "bar"
chart2.title = "Top 10 Products by Revenue"
chart2.y_axis.title = "Revenue (Rs.)"
chart2.height = 14
chart2.width  = 22

n_top = min(10, len(products_sum))
data2 = Reference(pws, min_col=4, min_row=2, max_row=n_top+2)
cats2 = Reference(pws, min_col=1, min_row=3, max_row=n_top+2)
chart2.add_data(data2, titles_from_data=True)
chart2.set_categories(cats2)
chart2.series[0].graphicalProperties.solidFill = MID_BLUE
pws.add_chart(chart2, "H3")

for col, w in [("A",32),("B",10),("C",12),("D",16),("E",16),("F",10)]:
    pws.column_dimensions[col].width = w
pws.freeze_panes = "A3"

# ═══════════════════════════════════════════════════════════════════════
# SHEET 5 – CUSTOMER ANALYSIS
# ═══════════════════════════════════════════════════════════════════════
cws = wb.create_sheet("Customer Analysis")
cws.sheet_view.showGridLines = False

cws.merge_cells("A1:E1")
cws["A1"] = "Top 20 Customers by Revenue"
cws["A1"].font = hdr_font(sz=13)
cws["A1"].fill = hfill(DARK_BLUE)
cws["A1"].alignment = Alignment(horizontal="center", vertical="center")
cws.row_dimensions[1].height = 28

write_header_row(cws, 2, ["Customer","Type","City","Orders","Revenue (Rs.)"])

for r, row in customers_sum.iterrows():
    dr = r + 3
    for ci, val in enumerate([row["Customer"],row["Type"],row["City"],row["Orders"],row["Revenue"]], 1):
        cell = cws.cell(dr, ci, val)
        cell.font = body_font(sz=9)
        cell.fill = hfill(LIGHT_GREY) if r % 2 == 0 else hfill(WHITE)
        cell.border = border()
        if ci == 5:
            cell.number_format = '#,##0'
    # Color-code customer type
    type_cell = cws.cell(dr, 2)
    if row["Type"] == "Wholesale":
        type_cell.font = Font(name="Arial", bold=True, size=9, color=MID_BLUE)
    else:
        type_cell.font = Font(name="Arial", bold=True, size=9, color="C0504D")

for col, w in [("A",35),("B",12),("C",14),("D",9),("E",16)]:
    cws.column_dimensions[col].width = w
cws.freeze_panes = "A3"

# ═══════════════════════════════════════════════════════════════════════
# SHEET 6 – REGION & EMPLOYEE
# ═══════════════════════════════════════════════════════════════════════
ews = wb.create_sheet("Region & Employee")
ews.sheet_view.showGridLines = False

ews.merge_cells("A1:E1")
ews["A1"] = "Sales Performance by Employee & Region"
ews["A1"].font = hdr_font(sz=13)
ews["A1"].fill = hfill(DARK_BLUE)
ews["A1"].alignment = Alignment(horizontal="center", vertical="center")
ews.row_dimensions[1].height = 28

write_header_row(ews, 2, ["Employee","Region","Orders","Revenue (Rs.)","Profit (Rs.)"])

for r, row in emp_sum.iterrows():
    dr = r + 3
    for ci, val in enumerate([row["Employee"],row["Region"],row["Orders"],row["Revenue"],row["Profit"]], 1):
        cell = ews.cell(dr, ci, val)
        cell.font = body_font(sz=10)
        cell.fill = hfill(LIGHT_GREY) if r % 2 == 0 else hfill(WHITE)
        cell.border = border()
        if ci in (4, 5):
            cell.number_format = '#,##0'

# Totals
tr2 = len(emp_sum) + 3
ews.cell(tr2, 1, "TOTAL").font = hdr_font(white=False)
ews.cell(tr2, 1).fill = hfill(LIGHT_BLUE)
for ci, col in [(3,"C"),(4,"D"),(5,"E")]:
    cell = ews.cell(tr2, ci, f"=SUM({col}3:{col}{tr2-1})")
    cell.font = hdr_font(white=False)
    cell.fill = hfill(LIGHT_BLUE)
    cell.border = border()
    if ci > 3:
        cell.number_format = '#,##0'

# Bar chart – Revenue by Employee
chart3 = BarChart()
chart3.type  = "col"
chart3.title = "Revenue by Sales Rep"
chart3.y_axis.title = "Revenue (Rs.)"
chart3.height = 13
chart3.width  = 20

data3  = Reference(ews, min_col=4, min_row=2, max_row=len(emp_sum)+2)
cats3  = Reference(ews, min_col=1, min_row=3, max_row=len(emp_sum)+2)
chart3.add_data(data3, titles_from_data=True)
chart3.set_categories(cats3)
chart3.series[0].graphicalProperties.solidFill = MID_BLUE
ews.add_chart(chart3, "G3")

for col, w in [("A",22),("B",12),("C",10),("D",16),("E",16)]:
    ews.column_dimensions[col].width = w
ews.freeze_panes = "A3"

# ── Save ───────────────────────────────────────────────────────────────
out_path = "/home/claude/lamination_project/excel/Lamination_Store_Analytics.xlsx"
wb.save(out_path)
print(f"Excel workbook saved: {out_path}")
print(f"Sheets: {[s.title for s in wb.worksheets]}")
